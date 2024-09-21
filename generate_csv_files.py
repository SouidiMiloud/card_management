import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, date

# Define headers and sample data types for each model based on the fields

def generate_sample_data(headers, sample_generators, num_records=5):
    """
    Generate sample data using provided headers and sample generators.
    """
    data = []
    for i in range(num_records):
        record = [gen(i) for gen in sample_generators]
        data.append(record)
    return data

def create_excel_file(filename, headers, data):
    """
    Create and save an Excel file with the given filename, headers, and data.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'

    # Write headers
    for col_num, header in enumerate(headers, start=1):
        sheet[f'{get_column_letter(col_num)}1'] = header

    # Write data
    for row_num, record in enumerate(data, start=2):
        for col_num, value in enumerate(record, start=1):
            sheet[f'{get_column_letter(col_num)}{row_num}'] = value

    # Save file
    workbook.save(filename)
    print(f'Excel file "{filename}" created successfully with {len(data)} records.')

# Define headers and sample data generators for each model

# SelfAccount Model
self_account_headers = [
    'country', 'description', 'price', 'account_type', 
    'account_url', 'proof', 'username', 'password'
]
self_account_generators = [
    lambda i: 'US',  # country
    lambda i: f'SelfAccount description {i+1}',  # description
    lambda i: round(i * 100.50, 2),  # price
    lambda i: f'AccountType {i+1}',  # account_type
    lambda i: f'http://account_url_{i+1}.com',  # account_url
    lambda i: f'image_{i+1}.png',  # proof
    lambda i: f'username_{i+1}',  # username
    lambda i: f'password_{i+1}',  # password
]

# HQCard Model
hq_card_headers = [
    'country', 'description', 'price', 'exp_date', 'name', 
    'address', 'city', 'state', 'email', 'phone', 'dob', 'ssn', 
    'ip', 'bank_name', 'base', 'card_bin', 'status', 'level', 
    'credit_debit', 'card_zip', 'ua', 'refundable'
]
hq_card_generators = [
    lambda i: 'US',  # country
    lambda i: f'HQCard description {i+1}',  # description
    lambda i: round(i * 150.75, 2),  # price
    lambda i: '12/25',  # exp_date
    lambda i: f'Name {i+1}',  # name
    lambda i: f'Address {i+1}',  # address
    lambda i: f'City {i+1}',  # city
    lambda i: 'State',  # state
    lambda i: f'email{i+1}@example.com',  # email
    lambda i: f'123-456-789{i}',  # phone
    lambda i: date(1990+i, 1, 1),  # dob
    lambda i: f'SSN{i+1}',  # ssn
    lambda i: f'192.168.0.{i+1}',  # ip
    lambda i: f'Bank {i+1}',  # bank_name
    lambda i: f'Base {i+1}',  # base
    lambda i: f'BIN{i+1}',  # card_bin
    lambda i: 'visa',  # status
    lambda i: 'standard',  # level
    lambda i: 'credit',  # credit_debit
    lambda i: f'ZIP{i+1}',  # card_zip
    lambda i: f'UserAgent {i+1}',  # ua
    lambda i: True  # refundable
]

# VHQCard Model
vhq_card_headers = hq_card_headers + ['months_left', 'screen_resolution', 'user_agent']
vhq_card_generators = hq_card_generators + [
    lambda i: i+1,  # months_left
    lambda i: '1920x1080',  # screen_resolution
    lambda i: f'UserAgentDetail {i+1}'  # user_agent
]

# WholesaleCard Model
wholesale_card_headers = [
    'country', 'description', 'price', 'exp_date', 'name', 'address', 
    'city', 'state', 'email', 'phone', 'dob', 'ssn', 'ip', 'number', 
    'quantity', 'user_agents', 'cvv'
]
wholesale_card_generators = [
    lambda i: 'US',  # country
    lambda i: f'WholesaleCard description {i+1}',  # description
    lambda i: round(i * 200.25, 2),  # price
    lambda i: '11/24',  # exp_date
    lambda i: f'Name {i+1}',  # name
    lambda i: f'Address {i+1}',  # address
    lambda i: f'City {i+1}',  # city
    lambda i: 'State',  # state
    lambda i: f'email{i+1}@example.com',  # email
    lambda i: f'123-456-789{i}',  # phone
    lambda i: date(1995+i, 1, 1),  # dob
    lambda i: f'SSN{i+1}',  # ssn
    lambda i: f'192.168.1.{i+1}',  # ip
    lambda i: f'123456781234567{i}',  # number
    lambda i: i+1,  # quantity
    lambda i: f'UserAgents {i+1}',  # user_agents
    lambda i: '123'  # cvv
]

# DumpCard Model
dump_card_headers = hq_card_headers + ['track1', 'track2']
dump_card_generators = hq_card_generators + [
    lambda i: f'Track1_{i+1}',  # track1
    lambda i: f'Track2_{i+1}'   # track2
]

# Generate data and create Excel files
self_account_data = generate_sample_data(self_account_headers, self_account_generators)
create_excel_file('self_account.xlsx', self_account_headers, self_account_data)

hq_card_data = generate_sample_data(hq_card_headers, hq_card_generators)
create_excel_file('hq_card.xlsx', hq_card_headers, hq_card_data)

vhq_card_data = generate_sample_data(vhq_card_headers, vhq_card_generators)
create_excel_file('vhq_card.xlsx', vhq_card_headers, vhq_card_data)

wholesale_card_data = generate_sample_data(wholesale_card_headers, wholesale_card_generators)
create_excel_file('wholesale_card.xlsx', wholesale_card_headers, wholesale_card_data)

dump_card_data = generate_sample_data(dump_card_headers, dump_card_generators)
create_excel_file('dump_card.xlsx', dump_card_headers, dump_card_data)
